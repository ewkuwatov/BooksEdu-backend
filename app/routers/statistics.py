from fastapi import APIRouter, Depends
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

from app.db.session import get_db
from app.models.university import University
from app.models.direction import Direction
from app.models.subject import Subject
from app.dependencies import require_owner_or_superadmin


router = APIRouter(prefix="/statistics", tags=["statistics"])


# ================= STYLES =================
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
title_font = Font(size=15, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ================= HELPERS =================
def calc_row_height(text: str, base=32, per_line=18):
    if not text:
        return base
    return base + text.count("\n") * per_line


def apply_column_widths(ws):
    widths = {
        1: 6,
        2: 70,
        3: 16,
        4: 30,
        5: 45,
        6: 22,
        7: 30,
        8: 26,
        9: 12,
        10: 12,
        11: 10,
        12: 14,
        13: 22,
        14: 18
    }

    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width


# ================= RENDER =================
def render_university(ws, uni, start_row=1):
    row = int(start_row)

    # ===== TITLE =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row, 1).value = uni.name
    ws.cell(row, 1).font = title_font
    ws.cell(row, 1).alignment = CENTER
    row += 2

    # ===== HEADER =====
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=13)
    ws.cell(row, 12).value = "ARMda mavjud shakli"
    ws.cell(row, 12).font = header_font
    ws.cell(row, 12).alignment = CENTER
    ws.cell(row, 12).fill = header_fill
    ws.cell(row, 12).border = border

    top_headers = [
        "№",
        "Мutaxassislik shifri va nomi",
        "Talabalar soni",
        "Fan nomi",
        "Adabiyot nomi",
        "Turi",
        "Muallif",
        "Nashriyot",
        "Til",
        "Yozuvi",
        "Yili",
    ]

    for col, text in enumerate(top_headers, 1):
        cell = ws.cell(row, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = CENTER
        cell.border = border

    ws.cell(row, 14).value = "Ta’minlanganlik %"
    ws.cell(row, 14).font = header_font
    ws.cell(row, 14).fill = header_fill
    ws.cell(row, 14).alignment = CENTER
    ws.cell(row, 14).border = border

    row += 1

    sub_headers = [""] * 11 + ["Elektron", "Bosma", ""]

    for col, text in enumerate(sub_headers, 1):
        cell = ws.cell(row, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = CENTER
        cell.border = border

    for col in range(1, 12):
        ws.merge_cells(start_row=row - 1, end_row=row, start_column=col, end_column=col)

    ws.merge_cells(start_row=row - 1, end_row=row, start_column=14, end_column=14)

    row += 1

    # ================= DATA =================
    subject_map = {}

    for d in uni.directions:
        for s in d.subjects:
            subject_map.setdefault(s.id, {
                "subject": s.name,
                "directions": set(),
                "students": 0,
                "literature": {}
            })

            block = subject_map[s.id]
            block["directions"].add(f"{d.number} - {d.name} ({d.course}-kurs)")
            block["students"] += d.student_count or 0

            for l in s.literature:
                block["literature"][l.id] = l

    index = 1

    for block in subject_map.values():
        start = row
        directions = "\n".join(sorted(block["directions"]))

        for i, lit in enumerate(block["literature"].values()):
            percent = 100 if lit.file_path else min(
                int((lit.printed_count or 0) * 6 / max(block["students"], 1) * 100),
                100
            )

            ws.append([
                index if i == 0 else "",
                directions if i == 0 else "",
                block["students"] if i == 0 else "",
                block["subject"] if i == 0 else "",
                lit.title,
                lit.kind,
                lit.author or "",
                lit.publisher or "",
                lit.language,
                lit.font_type,
                lit.year,
                "Mavjud" if lit.file_path else "",
                lit.printed_count or 0,
                f"{percent}%"
            ])

            for c in range(1, 15):
                ws.cell(row, c).alignment = CENTER
                ws.cell(row, c).border = border

            ws.row_dimensions[row].height = max(
                35,
                calc_row_height(directions) if i == 0 else 35
            )

            row += 1

        # 🔥 FIX — защита от row < start
        if row - 1 >= start:
            for col in [1, 2, 3, 4]:
                ws.merge_cells(
                    start_row=start,
                    end_row=row - 1,
                    start_column=col,
                    end_column=col
                )

        index += 1

    apply_column_widths(ws)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.sheet_view.zoomScale = 80

    return row + 2


# ================= ALL UNIVERSITIES =================
def render_all_universities(ws, universities):
    row = 1
    for uni in universities:
        row = render_university(ws, uni, start_row=row)


# ================= ROUTE =================
@router.get("/export")
async def export_statistics(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_owner_or_superadmin)
):
    query = select(University).options(
        selectinload(University.directions)
        .selectinload(Direction.subjects)
        .selectinload(Subject.literature)
    )

    universities = (await db.execute(query)).scalars().all()

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Umumiy"

    render_all_universities(ws_all, universities)

    for uni in universities:
        ws = wb.create_sheet(title=uni.name[:31])
        render_university(ws, uni)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=statistics.xlsx"}
    )
